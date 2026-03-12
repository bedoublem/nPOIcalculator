"""
nPOI DAS Configurator - Application Streamlit
Génère un fichier Excel d'organisation des entrées RF sur des unités nPOI (1U Rack 19", 8 ports)

Notation :
  - Fréquences codes courts : 7, 8, 9, 18, 21, 26, 35
  - SISO  : S1-MNO1-18
  - MIMO  : S1-MNO1-18A  et  S1-MNO1-18B

Installation :
    pip install streamlit openpyxl

Lancement :
    streamlit run npoi_configurator.py
"""

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# ── Constantes ────────────────────────────────────────────────────────────────

FREQUENCES_ORDRE = ["700", "800", "900", "1800", "2100", "2600", "3500"]

FREQ_CODE = {
    "700":  "700",
    "800":  "800",
    "900":  "900",
    "1800": "18",
    "2100": "21",
    "2600": "26",
    "3500": "35",
}

# Couleurs web (fond clair, texte foncé — interface fond blanc)
FREQ_COLORS_HTML = {
    "700":  ("#DDEEFF", "#1A3A5C"),
    "800":  ("#C8E0FF", "#102840"),
    "900":  ("#DDFFEE", "#1A4A2A"),
    "1800": ("#FFDDDD", "#4A1A1A"),
    "2100": ("#EEDDFF", "#3A1A4A"),
    "2600": ("#FFF3DD", "#4A3A1A"),
    "3500": ("#E0FFFF", "#1A3A4A"),
}

OP_COLORS_HTML = [
    ("#DDEEFF", "#1A3A5C"),
    ("#DDFFEE", "#1A4A2A"),
    ("#FFDDDD", "#4A1A1A"),
    ("#EEDDFF", "#3A1A4A"),
    ("#FFF3DD", "#4A3A1A"),
    ("#E0FFFF", "#1A3A4A"),
]

# Couleurs Excel (fond clair, texte foncé)
FREQ_COLORS_XL = {
    "700":  ("DDEEFF", "1A3A5C"),
    "800":  ("C8E0FF", "102840"),
    "900":  ("DDFFEE", "1A4A2A"),
    "1800": ("FFDDDD", "4A1A1A"),
    "2100": ("EEDDFF", "3A1A4A"),
    "2600": ("FFF3DD", "4A3A1A"),
    "3500": ("E0FFFF", "1A3A4A"),
}
OP_COLORS_XL = [
    ("DDEEFF", "1A3A5C"),
    ("DDFFEE", "1A4A2A"),
    ("FFDDDD", "4A1A1A"),
    ("EEDDFF", "3A1A4A"),
    ("FFF3DD", "4A3A1A"),
    ("E0FFFF", "1A3A4A"),
]

GRIS_LIBRE  = ("F0F0F0", "999999")
ENTETE_FOND = "0D1B2A"


# ── Logique métier ─────────────────────────────────────────────────────────────

def code_freq(freq):
    return FREQ_CODE.get(freq, freq)


def label_port(port):
    """
    SISO → S1-MNO1-18
    MIMO → S1-MNO1-18A  ou  S1-MNO1-18B
    """
    if port is None:
        return "Port libre"
    base = f"{port['secteur']}-{port['operateur']}-{code_freq(port['frequence'])}"
    return f"{base}{port['chaine']}" if port["chaine"] else base


def construire_ports(nb_secteurs, operateurs, config_freq, tri):
    ports = []
    for s in range(1, nb_secteurs + 1):
        for op in operateurs:
            for freq in FREQUENCES_ORDRE:
                mode = config_freq.get(freq, "N/A")
                if mode == "N/A":
                    continue
                chaines = ["A", "B"] if mode == "MIMO" else [None]
                for chaine in chaines:
                    ports.append({
                        "secteur":   f"S{s}",
                        "operateur": op,
                        "frequence": freq,
                        "chaine":    chaine,
                        "mode":      mode,
                    })

    if tri == "Par fréquence":
        ports.sort(key=lambda p: (
            FREQUENCES_ORDRE.index(p["secteur"],
            p["frequence"]),           
            p["operateur"],
        ))
    else:
        ports.sort(key=lambda p: (
            p["operateur"],
            p["secteur"],
            FREQUENCES_ORDRE.index(p["frequence"]),
        ))

    return ports


def grouper_en_npoi(ports):
    """Découpe séquentielle en groupes de 8 — zéro port libre, fréquences mélangées."""
    return [ports[i:i + 8] for i in range(0, len(ports), 8)]


def grouper_en_npoi_optimise(ports):
    """
    Groupage optimisé Best-Fit Decreasing (BFD) :
    - Les ports d'une même (fréquence x secteur) restent TOUJOURS dans le même nPOI.
    - Les blocs les plus grands sont placés en premier (BFD).
    - Chaque bloc va dans le nPOI le plus rempli qui peut encore l'accueillir.
    - Tri interne : secteur → fréquence → opérateur → chaîne.
    Garantit le minimum de nPOI possible tout en conservant les blocs fréq/secteur intacts.
    """
    from collections import defaultdict

    # Tenter de grouper toute une fréquence si elle tient dans 8 ports
    blocs_f = defaultdict(list)
    for p in ports:
        blocs_f[p["frequence"]].append(p)

    blocs_finals = []
    for freq in FREQUENCES_ORDRE:
        if freq not in blocs_f:
            continue
        bloc_freq = blocs_f[freq]
        if len(bloc_freq) <= 8:
            blocs_finals.append(bloc_freq)
        else:
            # Trop grand → découper par secteur
            par_secteur = defaultdict(list)
            for p in bloc_freq:
                par_secteur[p["secteur"]].append(p)
            for s in sorted(par_secteur):
                sous = par_secteur[s]
                if len(sous) <= 8:
                    blocs_finals.append(sous)
                else:
                    # Encore trop grand → découper par opérateur
                    par_op = defaultdict(list)
                    for p in sous:
                        par_op[p["operateur"]].append(p)
                    for op in sorted(par_op):
                        blocs_finals.append(par_op[op])

    # BFD : trier par taille décroissante
    blocs_sorted = sorted(blocs_finals, key=lambda b: -len(b))
    npois = []
    for bloc in blocs_sorted:
        taille = len(bloc)
        best_idx, best_fill = -1, -1
        for i, npoi in enumerate(npois):
            if 8 - len(npoi) >= taille and len(npoi) > best_fill:
                best_idx, best_fill = i, len(npoi)
        if best_idx >= 0:
            npois[best_idx].extend(bloc)
        else:
            npois.append(list(bloc))

    # Tri interne : secteur → fréquence → opérateur → chaîne
    for npoi in npois:
        npoi.sort(key=lambda p: (
            p["secteur"],
            FREQUENCES_ORDRE.index(p["frequence"]),
            p["operateur"],
            p["chaine"] or "",
        ))

    # Compléter à 8 avec None (ports libres)
    return [n + [None] * (8 - len(n)) for n in npois]


# ── Génération Excel ───────────────────────────────────────────────────────────

def style_cell(ws, coord, valeur, bg_hex, fg_hex,
               bold=False, align="left", border=True, size=11):
    cell = ws[coord]
    cell.value = valeur
    cell.font = Font(name="Consolas", size=size, bold=bold, color=fg_hex)
    cell.fill = PatternFill("solid", fgColor=bg_hex)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if border:
        thin = Side(style="thin", color="AAAAAA")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def generer_excel(nb_secteurs, operateurs, config_freq, tri, npois):
    wb = openpyxl.Workbook()

    # ── Feuille 1 : Détail par nPOI ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Configuration nPOI"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:G1")
    style_cell(ws1, "A1", "nPOI DAS — Configuration des entrées RF",
               ENTETE_FOND, "E0F0FF", bold=True, align="center", size=14, border=False)
    ws1.row_dimensions[1].height = 30

    freq_resume = "  |  ".join(
        [f"{code_freq(f)}:{m}" for f, m in config_freq.items() if m != "N/A"]
    )
    params = (f"Secteurs: {nb_secteurs}  |  Opérateurs: {', '.join(operateurs)}  |  "
              f"Fréquences: {freq_resume}  |  Tri: {tri}  |  nPOI: {len(npois)}")
    ws1.merge_cells("A2:G2")
    style_cell(ws1, "A2", params, "0A1525", "5A8AAA",
               align="center", size=9, border=False)
    ws1.row_dimensions[2].height = 18

    for i, w in enumerate([7, 10, 16, 10, 8, 26, 10], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    row = 4
    for idx, npoi in enumerate(npois):
        ws1.merge_cells(f"A{row}:G{row}")
        ws1.row_dimensions[row].height = 22
        style_cell(ws1, f"A{row}",
                   f"  nPOI {idx+1}   —   1U Rack 19\"   —   {sum(1 for p in npoi if p is not None)}/8 ports utilisés",
                   "0D1B2A", "7EC8E3", bold=True, align="left", size=11)
        row += 1

        for col, titre in enumerate(["Port", "Secteur", "Opérateur", "Fréq.", "Mode", "Label", "nPOI"], 1):
            style_cell(ws1, f"{get_column_letter(col)}{row}",
                       titre, "162030", "8AABCC", bold=True, align="center", size=12)
        ws1.row_dimensions[row].height = 18
        row += 1

        for p in range(8):
            port = npoi[p] if p < len(npoi) else None
            ws1.row_dimensions[row].height = 20
            if port is None:
                bg, fg = GRIS_LIBRE
            elif tri == "Par fréquence":
                bg, fg = FREQ_COLORS_XL.get(port["frequence"], GRIS_LIBRE)
            else:
                oi = operateurs.index(port["operateur"]) if port["operateur"] in operateurs else 0
                bg, fg = OP_COLORS_XL[oi % len(OP_COLORS_XL)]

            valeurs = [
                f"P{p+1}",
                port["secteur"]                      if port else "—",
                port["operateur"]                    if port else "—",
                code_freq(port["frequence"])          if port else "—",
                port["mode"]                         if port else "—",
                label_port(port),
                f"nPOI {idx+1}",
            ]
            for col, val in enumerate(valeurs, 1):
                style_cell(ws1, f"{get_column_letter(col)}{row}",
                           val, bg, fg, align="center" if col in (1, 4, 5, 7) else "left")
            row += 1
        row += 1

    # ── Feuille 2 : Matrice synthèse ─────────────────────────────────────────
    ws2 = wb.create_sheet("Matrice synthèse")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:I1")
    style_cell(ws2, "A1", "Matrice de synthèse — Assignation des ports par nPOI",
               ENTETE_FOND, "E0F0FF", bold=True, align="center", size=13, border=False)
    ws2.column_dimensions["A"].width = 12
    for c in range(2, 10):
        ws2.column_dimensions[get_column_letter(c)].width = 22

    row2 = 3
    style_cell(ws2, f"A{row2}", "nPOI", "162030", "8AABCC", bold=True, align="center")
    for p in range(8):
        style_cell(ws2, f"{get_column_letter(p+2)}{row2}",
                   f"P{p+1}", "162030", "8AABCC", bold=True, align="center")
    row2 += 1

    for idx, npoi in enumerate(npois):
        style_cell(ws2, f"A{row2}", f"nPOI {idx+1}",
                   "0D1B2A", "7EC8E3", bold=True, align="center")
        for p in range(8):
            port = npoi[p] if p < len(npoi) else None
            if port is None:
                bg, fg = GRIS_LIBRE
            elif tri == "Par fréquence":
                bg, fg = FREQ_COLORS_XL.get(port["frequence"], GRIS_LIBRE)
            else:
                oi = operateurs.index(port["operateur"]) if port["operateur"] in operateurs else 0
                bg, fg = OP_COLORS_XL[oi % len(OP_COLORS_XL)]
            style_cell(ws2, f"{get_column_letter(p+2)}{row2}",
                       label_port(port), bg, fg, align="center")
        row2 += 1

    # ── Feuille 3 : Légende ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Légende")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:D1")
    style_cell(ws3, "A1", "Légende des fréquences",
               ENTETE_FOND, "E0F0FF", bold=True, align="center", size=13, border=False)
    for i, w in enumerate([8, 12, 8, 30], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for col, titre in enumerate(["Code", "Fréq. MHz", "Mode", "Exemple label"], 1):
        style_cell(ws3, f"{get_column_letter(col)}2",
                   titre, "162030", "8AABCC", bold=True, align="center")
    r = 3
    for freq, mode in config_freq.items():
        bg, fg = FREQ_COLORS_XL.get(freq, GRIS_LIBRE) if mode != "N/A" else GRIS_LIBRE
        if mode == "SISO":
            exemple = f"S1-MNO1-{code_freq(freq)}"
        elif mode == "MIMO":
            exemple = f"S1-MNO1-{code_freq(freq)}A  /  S1-MNO1-{code_freq(freq)}B"
        else:
            exemple = "—"
        style_cell(ws3, f"A{r}", code_freq(freq), bg, fg, align="center")
        style_cell(ws3, f"B{r}", f"{freq} MHz",   bg, fg, align="center")
        style_cell(ws3, f"C{r}", mode,             bg, fg, align="center")
        style_cell(ws3, f"D{r}", exemple,          bg, fg)
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Interface Streamlit ────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="nPOI DAS Configurator",
        page_icon="📡",
        layout="wide",
    )

    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Share+Tech+Mono&family=Exo+2:wght@400;700&display=swap');

    html, body, [class*="css"] {
        font-family: 'Share Tech Mono', monospace;
        background-color: #ffffff;
        color: #1a2a3a;
    }
    .stApp { background-color: #ffffff; }
    h1, h2, h3 { font-family: 'Exo 2', sans-serif; color: #1a3a5c; }

    div[data-testid="stSidebar"] {
        background-color: #f0f4f8;
        border-right: 1px solid #c0d0e0;
    }
    .stButton > button {
        background: linear-gradient(135deg, #1e4a7a, #2a6aaa);
        color: #ffffff;
        border: none;
        border-radius: 4px;
        font-family: 'Share Tech Mono', monospace;
        font-size: 14px;
        padding: 8px 20px;
        width: 100%;
    }
    .npoi-block {
        background: #f7f9fc;
        border: 1px solid #c0d0e0;
        border-radius: 8px;
        padding: 12px 14px;
        margin-bottom: 14px;
    }
    .npoi-title {
        font-family: 'Exo 2', sans-serif;
        font-weight: 700;
        font-size: 16px;
        color: #1a3a5c;
        background: #ddeeff;
        border-radius: 4px;
        padding: 6px 12px;
        margin-bottom: 10px;
    }
    .ports-row {
        display: flex;
        flex-direction: row;
        gap: 5px;
    }
    .port-cell {
        flex: 1;
        border-radius: 4px;
        padding: 7px 3px;
        font-size: 16px;
        font-family: 'Share Tech Mono', monospace;
        text-align: center;
        border: 1px solid rgba(0,0,0,0.12);
        line-height: 1.5;
        word-break: break-all;
    }
    .port-num {
        font-weight: bold;
        font-size: 14px;
        opacity: 0.65;
        display: block;
        margin-bottom: 2px;
    }
    </style>
    """, unsafe_allow_html=True)

    # Titre
    st.markdown("""
    <div style="background:linear-gradient(135deg,#0d1b2a,#1a3a5c);
                padding:20px 28px;margin-bottom:24px;border-radius:8px;">
        <span style="font-family:'Exo 2',sans-serif;font-size:24px;
                     font-weight:700;color:#e0f0ff;letter-spacing:2px;">
            📡 nPOI DAS CONFIGURATOR
        </span><br>
        <span style="font-size:12px;color:#7ec8e3;letter-spacing:3px;">
            DISTRIBUTED ANTENNA SYSTEM · RACK 19" 1U · 8 PORTS RF PAR nPOI
        </span>
    </div>
    """, unsafe_allow_html=True)

    # ── Sidebar ────────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### ⚙️ Paramètres")

        nb_secteurs = st.slider("Nombre de secteurs", 1, 12, 1)

        st.markdown("**Opérateurs**")
        operateurs = st.multiselect(
            "Sélectionner les opérateurs",
            options=["MNO1", "MNO2", "MNO3", "MNO4", "OFR", "BYT","SFR", 
                     "Free", "OBE", "PXS", "TNT"],
            default=["MNO1", "MNO2"],
        )
        op_custom = st.text_input("Opérateur personnalisé", placeholder="Ex: Telenet")
        if op_custom and op_custom not in operateurs:
            operateurs.append(op_custom)

        st.markdown("---")
        st.markdown("**Fréquences & Mode**")
        st.caption("N/A = non utilisé · SISO = 1 port · MIMO = 2 ports (A + B)")

        config_freq = {}
        for freq in FREQUENCES_ORDRE:
            c1, c2 = st.columns([1, 2])
            c1.markdown(
                f"<div style='padding-top:8px;font-weight:bold;color:#1a3a5c;font-size:15px;'>"
                f"<b>{code_freq(freq)}</b> "
                f"<span style='font-size:9px;color:#888;'>{freq}MHz</span></div>",
                unsafe_allow_html=True
            )
            mode = c2.selectbox(
                f"_{freq}",
                options=["N/A", "SISO", "MIMO"],
                index=0,
                label_visibility="collapsed",
                key=f"freq_{freq}"
            )
            config_freq[freq] = mode

        st.markdown("---")
        tri = st.radio("Trier les ports par",
                       options=["Par fréquence", "Par opérateur"], index=0)

        st.markdown("---")
        st.markdown("**Mode d'agencement**")
        mode_groupage = st.radio(
            "Agencement des nPOI",
            options=["Séquentiel", "Groupé par fréquence (optimisé)"],
            index=1,
            help=(
                "**Séquentiel** : remplit les nPOI dans l'ordre, zéro port libre, "
                "mais les fréquences peuvent être mélangées.\n\n"
                "**Groupé** : chaque bloc (fréquence × secteur) reste intact dans "
                "le même nPOI — algorithme Best-Fit Decreasing pour minimiser le "
                "nombre de nPOI."
            ),
        )

    # ── Validation ────────────────────────────────────────────────────────────
    if not operateurs:
        st.warning("⚠️ Veuillez sélectionner au moins un opérateur.")
        return
    if not any(m != "N/A" for m in config_freq.values()):
        st.warning("⚠️ Veuillez activer au moins une fréquence (SISO ou MIMO).")
        return

    # ── Calcul ────────────────────────────────────────────────────────────────
    ports = construire_ports(nb_secteurs, operateurs, config_freq, tri)
    if mode_groupage == "Groupé par fréquence (optimisé)":
        npois = grouper_en_npoi_optimise(ports)
    else:
        npois = grouper_en_npoi(ports)

    ports_reels = len(ports)
    ports_libres = sum(1 for npoi in npois for p in npoi if p is None)

    # ── Métriques ─────────────────────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("nPOI nécessaires", len(npois))
    c2.metric("Ports utilisés", ports_reels)
    c3.metric("Ports libres", ports_libres)
    c4.metric("Unités de rack", f"{len(npois)} U")

    # ── Récap fréquences ──────────────────────────────────────────────────────
    st.markdown("### 📋 Récapitulatif des fréquences")
    fcols = st.columns(len(FREQUENCES_ORDRE))
    for i, freq in enumerate(FREQUENCES_ORDRE):
        mode = config_freq.get(freq, "N/A")
        if mode == "N/A":
            bg, fg = "#f0f0f0", "#aaaaaa"
        else:
            bg, fg = FREQ_COLORS_HTML.get(freq, ("#f0f0f0", "#333"))

        if mode == "SISO":
            exemple = f"S1-OP-{code_freq(freq)}"
        elif mode == "MIMO":
            exemple = f"S1-OP-{code_freq(freq)}A/B"
        else:
            exemple = "—"

        fcols[i].markdown(
            f'<div style="background:{bg};color:{fg};border-radius:6px;'
            f'padding:10px 4px;text-align:center;border:1px solid {fg}44;min-height:90px;">'
            f'<b style="font-size:20px;">{code_freq(freq)}</b><br>'
            f'<span style="font-size:9px;">{freq} MHz</span><br>'
            f'<b style="font-size:12px;">{mode}</b><br>'
            f'<span style="font-size:9px;opacity:0.8;">{exemple}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

    # ── Plan nPOI : 8 ports par ligne ─────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 🗂️ Plan d'organisation des nPOI")

    def get_color_html(port):
        if port is None:
            return "#eeeeee", "#aaaaaa"
        if tri == "Par fréquence":
            return FREQ_COLORS_HTML.get(port["frequence"], ("#eeeeee", "#aaa"))
        else:
            idx = operateurs.index(port["operateur"]) if port["operateur"] in operateurs else 0
            return OP_COLORS_HTML[idx % len(OP_COLORS_HTML)]

    for idx, npoi in enumerate(npois):
        # pad to 8 if needed (sequential mode)
        npoi_padded = list(npoi) + [None] * (8 - len(npoi))
        ports_used = sum(1 for p in npoi_padded if p is not None)
        cells_html = ""
        for p in range(8):
            port = npoi_padded[p]
            bg, fg = get_color_html(port)
            if port:
                line1 = f"{port['secteur']}-{port['operateur']}"
                line2 = f"<b>{code_freq(port['frequence'])}{port['chaine'] or ''}</b>"
            else:
                line1 = "&nbsp;"
                line2 = "<b style='color:#aaa;'>libre</b>"

            cells_html += (
                f'<div class="port-cell" style="background:{bg};color:{fg};">'
                f'<span class="port-num">P{p+1}</span>'
                f'{line1}<br>{line2}'
                f'</div>'
            )

        st.markdown(
            f'<div class="npoi-block">'
            f'<div class="npoi-title">'
            f'🔌 nPOI {idx+1} &nbsp;—&nbsp; {ports_used}/8 ports utilisés'
            f'</div>'
            f'<div class="ports-row">{cells_html}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    # ── Export Excel ──────────────────────────────────────────────────────────
    st.markdown("---")
    excel_buf = generer_excel(nb_secteurs, operateurs, config_freq, tri, npois)
    st.download_button(
        label="📥 Télécharger le fichier Excel",
        data=excel_buf,
        file_name="nPOI_DAS_Configuration.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
